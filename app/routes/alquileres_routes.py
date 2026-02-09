# ==========================================
# RUTAS FLASK PARA SISTEMA DE ALQUILERES
# ==========================================

"""
AGREGAR ESTAS RUTAS A TU APLICACIÃ“N FLASK
Crear archivo: app/routes/alquileres_routes.py
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from sqlalchemy import func, and_, or_
from app import db
from app.models import (
    PorcentajeGanancia, SemanaAlquiler, DetalleAlquilerSemanal,
    Alquiler, Vehiculo, Inquilino, Propietario, Banco, Usuario, EstadoAlquiler,
    TrabajoVehiculo, TipoTrabajo, Mecanico, Contrato, Deposito, ConfiguracionAlquiler
)
from functools import wraps
import os
from werkzeug.utils import secure_filename

alquileres_bp = Blueprint('alquiler', __name__)

# ==========================================
# DECORATOR PARA ADMIN
# ==========================================
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'admin':
            return jsonify({
                'success': False, 
                'message': 'Acceso denegado. Se requieren privilegios de administrador.'
            }), 403
        return f(*args, **kwargs)
    return decorated_function

# ==========================================
# PANTALLA PRINCIPAL - ALQUILERES
# ==========================================
@alquileres_bp.route('/')
@login_required
def index():
    """Pantalla principal de gestión de alquileres semanales"""
    
    # Get semanas
    semanas = SemanaAlquiler.query.order_by(
        SemanaAlquiler.fecha_inicio.desc()
    ).all()
    
    # Get porcentajes activos
    porcentajes_activos = PorcentajeGanancia.query.filter_by(activo=True).all()
    
    # Get bancos
    bancos = Banco.query.all()
    
    # Get tipos de trabajo para inversiones
    tipos_trabajo = TipoTrabajo.query.all()
    
    # Get mecánicos activos
    mecanicos = Mecanico.query.filter_by(activo=True).all()
    
    # CORRECCIÓN: Convertir a diccionarios para JSON
    mecanicos_data = [
        {
            'id': m.id,
            'nombre': m.nombre,
            'especialidad': m.especialidad or ''
        }
        for m in mecanicos
    ]
    
    tipos_trabajo_data = [
        {
            'id': t.id,
            'nombre': t.nombre,
            'descripcion': t.descripcion or ''
        }
        for t in tipos_trabajo
    ]
    
    # Calculate stats
    total_semanas = SemanaAlquiler.query.count()
    semanas_activas = SemanaAlquiler.query.filter_by(estado='abierta').count()
    
    # Pagos pendientes
    pagos_pendientes = DetalleAlquilerSemanal.query.filter_by(
        pago_confirmado=False
    ).count()
    
    # Ingreso del mes actual
    mes_actual = date.today().month
    anio_actual = date.today().year
    ingreso_total_mes = db.session.query(
        func.sum(SemanaAlquiler.ingreso_total)
    ).filter(
        and_(
            func.extract('month', SemanaAlquiler.fecha_inicio) == mes_actual,
            func.extract('year', SemanaAlquiler.fecha_inicio) == anio_actual
        )
    ).scalar() or 0
    
    return render_template(
        'modulos/alquileres.html',
        semanas=semanas,
        porcentajes_activos=porcentajes_activos,
        bancos=bancos,
        tipos_trabajo=tipos_trabajo,  # Para los selects del template
        mecanicos=mecanicos,  # Para los selects del template
        mecanicos_json=mecanicos_data,  # ✅ Para JavaScript
        tipos_trabajo_json=tipos_trabajo_data,  # ✅ Para JavaScript
        total_semanas=total_semanas,
        semanas_activas=semanas_activas,
        pagos_pendientes=pagos_pendientes,
        ingreso_total_mes=ingreso_total_mes,
        user_rol=current_user.rol
    )

# ==========================================
# CREAR SEMANA (CORREGIDA)
# ==========================================
@alquileres_bp.route('/alquiler/semanas/crear__', methods=['POST'])
@login_required
def crear_semana__():
    """Crea una nueva semana de trabajo"""
    
    try:
        fecha_inicio = datetime.strptime(request.form.get('fecha_inicio'), '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(request.form.get('fecha_fin'), '%Y-%m-%d').date()
        porcentaje_ganancia_id = int(request.form.get('porcentaje_ganancia_id'))
        notas = request.form.get('notas')
        
        # Validate dates
        if fecha_fin < fecha_inicio:
            flash('La fecha de fin debe ser posterior a la fecha de inicio', 'error')
            return redirect(url_for('alquiler.index'))
        
        # Get numero de semana
        numero_semana = fecha_inicio.isocalendar()[1]
        anio = fecha_inicio.year
        
        # Check if semana already exists
        existing = SemanaAlquiler.query.filter(
            and_(
                SemanaAlquiler.fecha_inicio == fecha_inicio,
                SemanaAlquiler.fecha_fin == fecha_fin
            )
        ).first()
        
        if existing:
            flash('Ya existe una semana con este rango de fechas', 'error')
            return redirect(url_for('alquiler.index'))
        
        # ✅ CORRECCIÓN: Validar que no haya otra semana activa en el mismo rango
        # ELIMINADO POR SOLICITUD DEL USUARIO
        # semana_solapada = SemanaAlquiler.query.filter(
        #     and_(
        #         SemanaAlquiler.estado == 'abierta',
        #         or_(
        #             and_(
        #                 SemanaAlquiler.fecha_inicio <= fecha_inicio,
        #                 SemanaAlquiler.fecha_fin >= fecha_inicio
        #             ),
        #             and_(
        #                 SemanaAlquiler.fecha_inicio <= fecha_fin,
        #                 SemanaAlquiler.fecha_fin >= fecha_fin
        #             ),
        #             and_(
        #                 SemanaAlquiler.fecha_inicio >= fecha_inicio,
        #                 SemanaAlquiler.fecha_fin <= fecha_fin
        #             )
        #         )
        #     )
        # ).first()
            
        # if semana_solapada:
        #     flash(f'Ya existe una semana activa que se solapa con este rango: {semana_solapada.fecha_inicio.strftime("%d/%m/%Y")} - {semana_solapada.fecha_fin.strftime("%d/%m/%Y")}', 'error')
        #     return redirect(url_for('alquiler.index'))
        
        # Create semana
        semana = SemanaAlquiler(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            numero_semana=numero_semana,
            anio=anio,
            porcentaje_ganancia_id=porcentaje_ganancia_id,
            estado='abierta',
            notas=notas,
            usuario_registro_id=current_user.id
        )
        
        db.session.add(semana)
        db.session.flush()
        
        # CORRECCIÓN: Obtener alquileres activos SOLO en este rango
        alquileres_activos = Alquiler.query.filter(
            and_(
                Alquiler.fecha_alquiler_inicio <= fecha_fin,
                Alquiler.fecha_alquiler_fin >= fecha_inicio,
                # Validar que no estén ya en otra semana activa
                ~Alquiler.id.in_(
                    db.session.query(DetalleAlquilerSemanal.alquiler_id).filter(
                        DetalleAlquilerSemanal.semana_alquiler_id != semana.id
                    )
                )
            )
        ).all()
        
        # Get porcentaje
        porcentaje = PorcentajeGanancia.query.get(porcentaje_ganancia_id)
        
        # Calculate dias de trabajo
        dias_trabajo = (fecha_fin - fecha_inicio).days + 1
        
        # Get fecha limite (jueves de la semana)
        fecha_limite = fecha_inicio + timedelta(days=(3 - fecha_inicio.weekday()) % 7)
        
        # Create detalles
        total_vehiculos = 0
        socios = set()
        inquilinos = set()
        ingreso_total = 0
        
        for alquiler in alquileres_activos:
            vehiculo = Vehiculo.query.get(alquiler.vehiculo_id)
            inquilino = Inquilino.query.get(alquiler.inquilino_id)
            propietario = Propietario.query.get(vehiculo.propietario_id)
            
            # ✅ CORRECCIÓN: Calcular precio diario correctamente
            precio_semanal = float(vehiculo.precio_semanal)
            precio_diario = precio_semanal / 7
            
            # Calcular días reales trabajados en esta semana
            dias_trabajados_semana = min(
                (min(alquiler.fecha_alquiler_fin, fecha_fin) - max(alquiler.fecha_alquiler_inicio, fecha_inicio)).days + 1,
                7
            )
            
            ingreso_calculado = precio_diario * dias_trabajados_semana
            nomina_empresa = ingreso_calculado * (float(porcentaje.porcentaje) / 100)
            
            # Check if tiene deuda
            tiene_deuda = date.today() > fecha_limite
            
            detalle = DetalleAlquilerSemanal(
                semana_alquiler_id=semana.id,
                alquiler_id=alquiler.id,
                vehiculo_id=alquiler.vehiculo_id,
                inquilino_id=alquiler.inquilino_id,
                propietario_id=vehiculo.propietario_id,
                precio_semanal=precio_semanal,
                dias_trabajo=dias_trabajados_semana,
                ingreso_calculado=ingreso_calculado,
                porcentaje_empresa=porcentaje.porcentaje,
                nomina_empresa=nomina_empresa,
                tiene_deuda=tiene_deuda,
                fecha_limite_pago=fecha_limite,
                nomina_final=ingreso_calculado,
                usuario_registro_id=current_user.id
            )
            
            db.session.add(detalle)
            
            total_vehiculos += 1
            socios.add(propietario.id)
            inquilinos.add(inquilino.id)
            ingreso_total += ingreso_calculado
        
        # Update semana totals
        semana.total_vehiculos = total_vehiculos
        semana.total_socios = len(socios)
        semana.total_inquilinos = len(inquilinos)
        semana.ingreso_total = ingreso_total
        
        db.session.commit()
        
        flash(f'Semana creada exitosamente con {total_vehiculos} vehículos', 'success')
        return redirect(url_for('alquiler.index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear semana: {str(e)}', 'error')
        return redirect(url_for('alquiler.index'))

# ==========================================
# CREAR SEMANA (CORREGIDA - SIN AUTO-AGREGAR ALQUILERES)
# ==========================================
@alquileres_bp.route('/alquiler/semanas/crear', methods=['POST'])
@login_required
def crear_semana():
    """Crea una nueva semana de trabajo VACÍA"""
    
    try:
        fecha_inicio = datetime.strptime(request.form.get('fecha_inicio'), '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(request.form.get('fecha_fin'), '%Y-%m-%d').date()
        porcentaje_ganancia_id = int(request.form.get('porcentaje_ganancia_id'))
        notas = request.form.get('notas')
        
        # Validate dates
        if fecha_fin < fecha_inicio:
            flash('La fecha de fin debe ser posterior a la fecha de inicio', 'error')
            return redirect(url_for('alquiler.index'))
        
        # Get numero de semana
        numero_semana = fecha_inicio.isocalendar()[1]
        anio = fecha_inicio.year
        
        # Check if semana already exists
        existing = SemanaAlquiler.query.filter(
            and_(
                SemanaAlquiler.fecha_inicio == fecha_inicio,
                SemanaAlquiler.fecha_fin == fecha_fin
            )
        ).first()
        
        if existing:
            flash('Ya existe una semana con este rango de fechas', 'error')
            return redirect(url_for('alquiler.index'))
        
        # Validar que no haya otra semana activa en el mismo rango
        # ELIMINADO POR SOLICITUD DEL USUARIO para permitir múltiples semanas
        # semana_solapada = SemanaAlquiler.query.filter(
        #     and_(
        #         SemanaAlquiler.estado == 'abierta',
        #         or_(
        #             and_(
        #                 SemanaAlquiler.fecha_inicio <= fecha_inicio,
        #                 SemanaAlquiler.fecha_fin >= fecha_inicio
        #             ),
        #             and_(
        #                 SemanaAlquiler.fecha_inicio <= fecha_fin,
        #                 SemanaAlquiler.fecha_fin >= fecha_fin
        #             ),
        #             and_(
        #                 SemanaAlquiler.fecha_inicio >= fecha_inicio,
        #                 SemanaAlquiler.fecha_fin <= fecha_fin
        #             )
        #         )
        #     )
        # ).first()
            
        # if semana_solapada:
        #     flash(f'Ya existe una semana activa que se solapa: {semana_solapada.fecha_inicio.strftime("%d/%m/%Y")} - {semana_solapada.fecha_fin.strftime("%d/%m/%Y")}', 'error')
        #     return redirect(url_for('alquiler.index'))
        
        # ✅ CORRECCIÓN: Crear semana VACÍA (sin alquileres automáticos)
        semana = SemanaAlquiler(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            numero_semana=numero_semana,
            anio=anio,
            porcentaje_ganancia_id=porcentaje_ganancia_id,
            estado='abierta',
            notas=notas,
            # ✅ Inicializar en 0
            total_vehiculos=0,
            total_socios=0,
            total_inquilinos=0,
            ingreso_total=0,
            usuario_registro_id=current_user.id,
            usuario_actualizo_id=current_user.id
        )
        
        db.session.add(semana)
        db.session.commit()
        
        flash('Semana creada exitosamente (vacía, lista para agregar alquileres)', 'success')
        return redirect(url_for('alquiler.index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear semana: {str(e)}', 'error')
        return redirect(url_for('alquiler.index'))

# ==========================================
# VALIDAR SEMANAS ACTIVAS (NUEVA)
# ==========================================
@alquileres_bp.route('/alquiler/semanas/validar-activas')
@login_required
def validar_semanas_activas():
    """Valida si hay semanas activas fuera del rango actual"""
    
    today = date.today()
    semanas_activas = SemanaAlquiler.query.filter_by(estado='abierta').all()
    
    problemas = []
    
    for semana in semanas_activas:
        es_semana_actual = semana.fecha_inicio <= today <= semana.fecha_fin
        
        if not es_semana_actual:
            dias_diferencia = (today - semana.fecha_fin).days if today > semana.fecha_fin else (semana.fecha_inicio - today).days
            
            problemas.append({
                'id': semana.id,
                'fecha_inicio': semana.fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': semana.fecha_fin.strftime('%d/%m/%Y'),
                'numero_semana': semana.numero_semana,
                'dias_diferencia': dias_diferencia,
                'tipo': 'pasada' if today > semana.fecha_fin else 'futura'
            })
    
    return jsonify({
        'success': True,
        'tiene_problemas': len(problemas) > 0,
        'total_activas': len(semanas_activas),
        'problemas': problemas
    })
    
# ==========================================
# VER DETALLES DE SEMANA
# ==========================================
@alquileres_bp.route('/alquiler/semanas/<int:id>/detalles')
@login_required
def ver_detalles_semana(id):
    """Retorna los detalles de una semana en formato JSON"""
    
    try:
        semana = SemanaAlquiler.query.get_or_404(id)
        
        # FILTRADO CORRECTO: Solo detalles de ESTA semana
        detalles = DetalleAlquilerSemanal.query.filter_by(
            semana_alquiler_id=id
        ).all()
        
        detalles_data = []
        for detalle in detalles:
            try:
                vehiculo = Vehiculo.query.get(detalle.vehiculo_id)
                inquilino = Inquilino.query.get(detalle.inquilino_id)
                propietario = Propietario.query.get(detalle.propietario_id)
                
                # Get marca y modelo
                marca_modelo = vehiculo.marca_modelo if vehiculo else None
                
                # Get iniciales propietario
                iniciales = '??'
                propietario_nombre = ''
                if propietario:
                    propietario_nombre = propietario.nombre_apellido or ''
                    nombre_parts = propietario_nombre.split() if propietario_nombre else ['?']
                    iniciales = ''.join([p[0].upper() for p in nombre_parts[:2]])
                
                # Get datos del vehículo
                vehiculo_marca = ''
                vehiculo_modelo_str = ''
                vehiculo_placa = ''
                if vehiculo:
                    vehiculo_placa = vehiculo.placa or ''
                    if marca_modelo:
                        vehiculo_marca = marca_modelo.marca or ''
                        vehiculo_modelo_str = marca_modelo.modelo or ''
                
                # Get datos del inquilino
                inquilino_nombre = ''
                inquilino_telefono = ''
                if inquilino:
                    inquilino_nombre = inquilino.nombre_apellido or ''
                    inquilino_telefono = inquilino.telefono or ''
                
                # Calcular inversiones con debug
                inversiones_totales = db.session.query(
                    func.sum(TrabajoVehiculo.costo)
                ).filter(
                    and_(
                        TrabajoVehiculo.vehiculo_id == detalle.vehiculo_id,
                        TrabajoVehiculo.fecha_inicio >= semana.fecha_inicio,
                        TrabajoVehiculo.fecha_inicio <= semana.fecha_fin
                    )
                ).scalar() or 0
                
                # --- NEW LOGIC: Contract & Deposits ---
                contrato_deuda = 0
                depositos_deuda = 0
                depositos_estado = 'Pagado'
                
                # Buscar contrato activo para este vehiculo e inquilino
                contrato_activo = Contrato.query.filter(
                    Contrato.vehiculo_id == detalle.vehiculo_id,
                    Contrato.inquilino_id == detalle.inquilino_id,
                    Contrato.estado == 'activo'
                ).first()
                
                if contrato_activo:
                    # Deuda contrato:
                    # User Logic: 
                    # - 'efectivo' or 'transferencia' -> PAGADO (Deuda 0)
                    # - 'pendiente' (or anything else presumably) -> NO PAGADO (Deuda = monto)
                    
                    # Normalizamos a minusculas por si acaso
                    tipo_pago = str(contrato_activo.tipo_pago).lower() if contrato_activo.tipo_pago else 'pendiente'
                    
                    if tipo_pago in ['efectivo', 'transferencia']:
                        contrato_deuda = 0
                    else:
                        # Asumimos que 'pendiente' o cualquier otro estado no pagado tiene deuda
                        contrato_deuda = float(contrato_activo.monto_contrato)
                    
                    # Deuda depositos: sumar pendiente de todos los depositos del contrato
                    for dep in contrato_activo.depositos:
                        depositos_deuda += float(dep.monto_pendiente)
                
                if depositos_deuda > 0:
                    depositos_estado = 'Pendiente'
                # --------------------------------------
                
                # Debug: Contar cuántos trabajos encontró
                count_trabajos = db.session.query(
                    func.count(TrabajoVehiculo.id)
                ).filter(
                    and_(
                        TrabajoVehiculo.vehiculo_id == detalle.vehiculo_id,
                        TrabajoVehiculo.fecha_inicio >= semana.fecha_inicio,
                        TrabajoVehiculo.fecha_inicio <= semana.fecha_fin
                    )
                ).scalar() or 0
                
                print(f"   Vehículo {vehiculo_placa}: {count_trabajos} trabajos, Total: ${inversiones_totales}")
                    
                detalles_data.append({
                    'id': detalle.id,
                    'vehiculo_id': detalle.vehiculo_id,
                    'inquilino_id': detalle.inquilino_id,
                    'propietario_id': detalle.propietario_id,
                    'propietario_nombre': propietario_nombre,
                    'propietario_iniciales': iniciales,
                    'vehiculo_marca': vehiculo_marca,
                    'vehiculo_modelo': vehiculo_modelo_str,
                    'vehiculo_placa': vehiculo_placa,
                    'inquilino_nombre': inquilino_nombre,
                    'inquilino_telefono': inquilino_telefono,
                    'contrato_deuda': contrato_deuda,      # NEW
                    'depositos_deuda': depositos_deuda,    # NEW
                    'precio_semanal': float(detalle.precio_semanal),
                    'dias_trabajo': detalle.dias_trabajo,
                    'ingreso_calculado': float(detalle.ingreso_calculado),
                    'inversion_mecanica': float(detalle.inversion_mecanica or 0),
                    'inversiones_totales': float(inversiones_totales),
                    'concepto_inversion': detalle.concepto_inversion or '',
                    'monto_descuento': float(detalle.monto_descuento or 0),
                    'concepto_descuento': detalle.concepto_descuento or '',
                    'nomina_empresa': float(detalle.nomina_empresa),
                    'porcentaje_empresa': float(detalle.porcentaje_empresa),
                    'tiene_deuda': detalle.tiene_deuda,
                    'monto_deuda': float(detalle.monto_deuda or 0),
                    'nomina_final': float(detalle.nomina_final),
                    'banco_id': detalle.banco_id,
                    'banco_nombre': detalle.banco.banco if detalle.banco else ('EFECTIVO' if detalle.pago_confirmado else ''),
                    'fecha_confirmacion_pago': detalle.fecha_confirmacion_pago.isoformat() if detalle.fecha_confirmacion_pago else '',
                    'pago_confirmado': detalle.pago_confirmado,
                    'comprobante_pago_path': detalle.comprobante_pago_path,
                    'notas': detalle.notas or ''
                })
            except Exception as e:
                print(f"❌ Error procesando detalle {detalle.id}: {str(e)}")
                continue
        
        return jsonify({
            'success': True,
            'semana': {
                'id': semana.id,
                'fecha_inicio': semana.fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': semana.fecha_fin.strftime('%d/%m/%Y'),
                'total_vehiculos': semana.total_vehiculos,
                'total_socios': semana.total_socios,
                'total_inquilinos': semana.total_inquilinos,
                'ingreso_total': float(semana.ingreso_total)
            },
            'detalles': detalles_data
        })
        
    except Exception as e:
        print(f"❌ Error en ver_detalles_semana: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500



# ==========================================
# GUARDAR CAMBIOS EN DETALLES
# ==========================================
@alquileres_bp.route('/alquiler/semanas/<int:id>/guardar-cambios', methods=['POST'])
@login_required
def guardar_cambios_semana(id):
    """Guarda los cambios realizados en los detalles de la semana"""
    
    try:
        data = request.get_json()
        cambios = data.get('cambios', [])
        
        updated_count = 0
        
        for cambio in cambios:
            detalle = DetalleAlquilerSemanal.query.get(cambio['id'])
            if detalle and detalle.semana_alquiler_id == id:
                # Update fields
                detalle.precio_semanal = cambio.get('precio_semanal')
                detalle.dias_trabajo = cambio.get('dias_trabajo')
                detalle.inversion_mecanica = cambio.get('inversion_mecanica', 0)
                detalle.concepto_inversion = cambio.get('concepto_inversion')
                detalle.monto_descuento = cambio.get('monto_descuento', 0)
                detalle.concepto_descuento = cambio.get('concepto_descuento')
                detalle.monto_deuda = cambio.get('monto_deuda', 0)
                detalle.banco_id = cambio.get('banco_id') if cambio.get('banco_id') else None
                
                if cambio.get('fecha_confirmacion_pago'):
                    detalle.fecha_confirmacion_pago = datetime.strptime(
                        cambio.get('fecha_confirmacion_pago'), '%Y-%m-%d'
                    ).date()
                
                detalle.pago_confirmado = cambio.get('pago_confirmado', False)
                detalle.notas = cambio.get('notas')
                
                # Recalcular usando nueva formula: (Semanal / 7) * DT y redondeando a ENTERO
                precio_diario = float(detalle.precio_semanal) / 7.0
                detalle.ingreso_calculado = int(round(precio_diario * float(detalle.dias_trabajo)))
                
                # Nomina Calculada: (Ingreso * 0.9) - Inversion
                detalle.nomina_empresa = (detalle.ingreso_calculado * 0.9) - float(detalle.inversion_mecanica or 0)
                
                # Nomina Final (Nomina Base + Deuda)
                detalle.nomina_final = detalle.ingreso_calculado + float(detalle.monto_deuda or 0)
                
                detalle.usuario_actualizo_id = current_user.id
                detalle.fecha_hora_actualizo = datetime.utcnow()
                
                updated_count += 1
        
        # Recalculate semana totals
        semana = SemanaAlquiler.query.get(id)
        if semana:
            detalles = DetalleAlquilerSemanal.query.filter_by(semana_alquiler_id=id).all()
            semana.ingreso_total = sum(float(d.ingreso_calculado) for d in detalles)
            semana.usuario_actualizo_id = current_user.id
            semana.fecha_hora_actualizo = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{updated_count} detalles actualizados',
            'updated': updated_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==========================================
# ALQUILERES DISPONIBLES PARA AGREGAR
# ==========================================

@alquileres_bp.route('/alquiler/semanas/<int:id>/alquileres_disponibles')
@login_required
def alquileres_disponibles(id):
    """Retorna alquileres activos que no estÃ¡n en esta semana (VERSIÃ“N MEJORADA)"""
    
    try:
        semana = SemanaAlquiler.query.get_or_404(id)
        
        # Get alquileres ya en esta semana
        alquileres_en_semana = db.session.query(
            DetalleAlquilerSemanal.alquiler_id
        ).filter_by(semana_alquiler_id=id).all()
        
        alquileres_ids_en_semana = [a[0] for a in alquileres_en_semana]
        
        # Get alquileres activos en el rango de la semana
        alquileres_disponibles = Alquiler.query.filter(
            and_(
                Alquiler.fecha_alquiler_inicio <= semana.fecha_fin,
                Alquiler.fecha_alquiler_fin >= semana.fecha_inicio,
                ~Alquiler.id.in_(alquileres_ids_en_semana)
            )
        ).all()
        
        alquileres_data = []
        for alquiler in alquileres_disponibles:
            vehiculo = Vehiculo.query.get(alquiler.vehiculo_id)
            inquilino = Inquilino.query.get(alquiler.inquilino_id)
            
            if not vehiculo or not inquilino:
                continue
            
            # Get propietario
            propietario = Propietario.query.get(vehiculo.propietario_id) if vehiculo else None
            
            # Get marca y modelo
            marca_modelo = vehiculo.marca_modelo_vehiculo if vehiculo else None
            
            alquileres_data.append({
                'id': alquiler.id,
                'vehiculo_placa': vehiculo.placa,
                'vehiculo_marca': marca_modelo.marca if marca_modelo else 'N/A',
                'vehiculo_modelo': marca_modelo.modelo if marca_modelo else 'N/A',
                'inquilino_nombre': inquilino.nombre_apellido,
                'inquilino_telefono': inquilino.telefono if inquilino.telefono else '',
                'propietario_nombre': propietario.nombre_apellido if propietario else '',
                'precio_semanal': float(vehiculo.precio_semanal) if vehiculo else 0,
                'fecha_inicio': alquiler.fecha_alquiler_inicio.isoformat() if alquiler.fecha_alquiler_inicio else '',
                'fecha_fin': alquiler.fecha_alquiler_fin.isoformat() if alquiler.fecha_alquiler_fin else ''
            })
        
        # Sort by placa
        alquileres_data.sort(key=lambda x: x['vehiculo_placa'])
        
        return jsonify({
            'success': True,
            'alquileres': alquileres_data,
            'total': len(alquileres_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
# ==========================================
# AGREGAR ALQUILER A SEMANA (VERSIÃ“N CORREGIDA)
# ==========================================

@alquileres_bp.route('/alquiler/semanas/agregar_alquiler', methods=['POST'])
@login_required
def agregar_alquiler_a_semana():
    """Crea un nuevo detalle de alquiler para la semana (CON ACTUALIZACIÓN DE ESTADO)"""
    
    try:
        data = request.get_json()
        semana_id = int(data.get('semana_id'))
        vehiculo_id = int(data.get('vehiculo_id'))
        inquilino_id = int(data.get('inquilino_id'))
        dias_trabajo = int(data.get('dias_trabajo', 7))
        
        # Validaciones existentes...
        existing = DetalleAlquilerSemanal.query.filter_by(
            semana_alquiler_id=semana_id,
            vehiculo_id=vehiculo_id
        ).first()
        
        if existing:
            return jsonify({
                'success': False,
                'message': 'Este vehículo ya está en esta semana'
            }), 400
        
        existing_inquilino = DetalleAlquilerSemanal.query.filter_by(
            semana_alquiler_id=semana_id,
            inquilino_id=inquilino_id
        ).first()
        
        if existing_inquilino:
            return jsonify({
                'success': False,
                'message': 'Este inquilino ya tiene un vehículo asignado en esta semana'
            }), 400
        
        # Get objetos
        semana = SemanaAlquiler.query.get_or_404(semana_id)
        vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
        
        # ✅ VERIFICAR que el vehículo esté disponible
        # VALIDACIÓN ELIMINADA: Se permite agregar alquileres a vehículos no disponibles
        # porque si tiene contrato activo, estará marcado como no disponible (ocupado)
        # if not vehiculo.disponible:
        #    return jsonify({
        #        'success': False,
        #        'message': 'El vehículo no está disponible para alquiler'
        #    }), 400
        
        estado = EstadoAlquiler.query.filter_by(nombre='activo').first()
        propietario = Propietario.query.get(vehiculo.propietario_id)
        
        if not propietario:
            return jsonify({
                'success': False,
                'message': 'El vehículo no tiene propietario asignado'
            }), 400
        
        inquilino = Inquilino.query.get_or_404(inquilino_id)
        
        if not estado:
            estado = EstadoAlquiler.query.first()
        
        # PASO 1: Crear registro en tabla ALQUILERES
        precio_semanal = float(vehiculo.precio_semanal)
        precio_diario = precio_semanal / 7
        ingreso = precio_diario * dias_trabajo
        
        nuevo_alquiler = Alquiler(
            vehiculo_id=vehiculo_id,
            inquilino_id=inquilino_id,
            estado_id=estado.id,
            fecha_alquiler_inicio=semana.fecha_inicio,
            fecha_alquiler_fin=semana.fecha_fin,
            semana=semana.numero_semana,
            dia_trabajo=dias_trabajo,
            ingreso=ingreso,
            monto_descuento=0.00,
            usuario_registro_id=current_user.id,
            usuario_actualizo_id=current_user.id
        )
        
        db.session.add(nuevo_alquiler)
        db.session.flush()
        
        
        # PASO 2: Crear el detalle
        porcentaje = PorcentajeGanancia.query.get(semana.porcentaje_ganancia_id)
        config = ConfiguracionAlquiler.query.first() # Get configuration
        
        ingreso_calculado = int(round(precio_diario * dias_trabajo))
        
        # Lavado Automático
        costo_lavado = 350.00
        if config and config.costo_lavado is not None:
             costo_lavado = float(config.costo_lavado)
             
        concepto_inversion = 'Lavado Automático al Ingreso'
        
        # Nomina Calculada: (Ingreso * 0.9) - Inversion (Lavado)
        # CORRECCION: La nomina empresa se calcula sobre el ingreso BRUTO generalmente...
        # Nomina Empresa = Ingreso * Porcentaje
        nomina_empresa = (ingreso_calculado * (porcentaje.porcentaje / 100))
        
        fecha_limite = semana.fecha_inicio + timedelta(
            days=(3 - semana.fecha_inicio.weekday()) % 7
        )
        
        detalle = DetalleAlquilerSemanal(
            semana_alquiler_id=semana_id,
            alquiler_id=nuevo_alquiler.id,
            vehiculo_id=vehiculo_id,
            inquilino_id=inquilino_id,
            propietario_id=vehiculo.propietario_id,
            precio_semanal=precio_semanal,
            dias_trabajo=dias_trabajo,
            ingreso_calculado=ingreso_calculado,
            porcentaje_empresa=porcentaje.porcentaje,
            nomina_empresa=nomina_empresa,
            fecha_limite_pago=fecha_limite,
            nomina_final=ingreso_calculado,
            usuario_registro_id=current_user.id,
            # Automatic Wash
            inversion_mecanica=0, 
            concepto_inversion=None 
        )
        
        db.session.add(detalle)
        db.session.flush() 
        
        # --- NEW: Create TrabajoVehiculo for Wash Cost ---
        if costo_lavado > 0:
            # Find a default mechanic (Try ID 1, otherwise first available)
            mecanico_defecto = Mecanico.query.get(1)
            if not mecanico_defecto:
                mecanico_defecto = Mecanico.query.filter_by(activo=True).first()
            
            # If still no mechanic, this will fail. Admin should ensure at least one mechanic exists.
            mecanico_id = mecanico_defecto.id if mecanico_defecto else None
            
            nuevo_lavado = TrabajoVehiculo(
                vehiculo_id=vehiculo_id,
                mecanico_id=mecanico_id, # Assign mechanic
                fecha_inicio=semana.fecha_inicio,
                descripcion='Lavado Automático al Ingreso',
                costo=costo_lavado,
                estado='completado',
                usuario_registro_id=current_user.id,
                usuario_actualizo_id=current_user.id
            )
            db.session.add(nuevo_lavado)
        # -------------------------------------------------
        
        # ✅ PASO 3: CAMBIAR ESTADO DEL VEHÍCULO A NO DISPONIBLE
        vehiculo.disponible = False
        vehiculo.usuario_actualizo_id = current_user.id
        vehiculo.fecha_hora_actualizo = datetime.utcnow()
        
        # Update semana totals
        semana.total_vehiculos = (semana.total_vehiculos or 0) + 1
        semana.ingreso_total = float(semana.ingreso_total or 0) + ingreso_calculado
        
        semana.total_socios = db.session.query(
            func.count(func.distinct(DetalleAlquilerSemanal.propietario_id))
        ).filter_by(semana_alquiler_id=semana_id).scalar() + 1
        
        semana.total_inquilinos = db.session.query(
            func.count(func.distinct(DetalleAlquilerSemanal.inquilino_id))
        ).filter_by(semana_alquiler_id=semana_id).scalar() + 1
        
        semana.usuario_actualizo_id = current_user.id
        semana.fecha_hora_actualizo = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Alquiler agregado exitosamente',
            'detalle_id': detalle.id,
            'alquiler_id': nuevo_alquiler.id
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        print(f"❌ Error al agregar alquiler: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@alquileres_bp.route('/alquiler/semanas/agregar_alquiler_', methods=['POST'])
@login_required
def agregar_alquiler_a_semana_():
    """Crea un nuevo detalle de alquiler para la semana"""
    
    #try:
    data = request.get_json()
    semana_id = int(data.get('semana_id'))
    vehiculo_id = int(data.get('vehiculo_id'))
    inquilino_id = int(data.get('inquilino_id'))
    dias_trabajo = int(data.get('dias_trabajo', 7))
    
    # Validar que no exista ya este vehÃ­culo en la semana
    existing = DetalleAlquilerSemanal.query.filter_by(
        semana_alquiler_id=semana_id,
        vehiculo_id=vehiculo_id
    ).first()
    
    if existing:
        return jsonify({
            'success': False,
            'message': 'Este vehÃ­culo ya estÃ¡ en esta semana'
        }), 400
    
    # Validar que no exista ya este inquilino en la semana
    existing_inquilino = DetalleAlquilerSemanal.query.filter_by(
        semana_alquiler_id=semana_id,
        inquilino_id=inquilino_id
    ).first()
    
    if existing_inquilino:
        return jsonify({
            'success': False,
            'message': 'Este inquilino ya tiene un vehÃ­culo asignado en esta semana'
        }), 400
    
    # Get semana
    semana = SemanaAlquiler.query.get_or_404(semana_id)
    
    # Get vehÃ­culo
    vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
    propietario = Propietario.query.get(vehiculo.propietario_id)
    
    if not propietario:
        return jsonify({
            'success': False,
            'message': 'El vehÃ­culo no tiene propietario asignado'
        }), 400
    
    # Get inquilino
    inquilino = Inquilino.query.get_or_404(inquilino_id)
    
    # Get porcentaje
    porcentaje = PorcentajeGanancia.query.get(semana.porcentaje_ganancia_id)
    
    # Calculate
    precio_semanal = float(vehiculo.precio_semanal)
    ingreso_calculado = int(round((precio_semanal / 7.0) * dias_trabajo))
    
    # Nomina Calculada: (Ingreso * 0.9) - Inversion (aquÃ­ inversion es 0 al crear)
    nomina_empresa = (ingreso_calculado * 0.9) - 0
    
    # Calculate fecha limite (jueves)
    fecha_limite = semana.fecha_inicio + timedelta(
        days=(3 - semana.fecha_inicio.weekday()) % 7
    )
    tiene_deuda = date.today() > fecha_limite
    
    # Create detalle
    detalle = DetalleAlquilerSemanal(
        semana_alquiler_id=semana_id,
        alquiler_id=None,  # No hay alquiler previo
        vehiculo_id=vehiculo_id,
        inquilino_id=inquilino_id,
        propietario_id=vehiculo.propietario_id,
        precio_semanal=precio_semanal,
        dias_trabajo=dias_trabajo,
        ingreso_calculado=ingreso_calculado,
        porcentaje_empresa=porcentaje.porcentaje,
        nomina_empresa=nomina_empresa,
        tiene_deuda=tiene_deuda,
        fecha_limite_pago=fecha_limite,
        nomina_final=ingreso_calculado,
        usuario_registro_id=current_user.id
    )
    
    db.session.add(detalle)
    
    # Update semana totals
    semana.total_vehiculos = (semana.total_vehiculos or 0) + 1
    semana.ingreso_total = float(semana.ingreso_total or 0) + ingreso_calculado
    
    # Recalculate unique propietarios and inquilinos
    semana.total_socios = db.session.query(
        func.count(func.distinct(DetalleAlquilerSemanal.propietario_id))
    ).filter_by(semana_alquiler_id=semana_id).scalar() + 1
    
    semana.total_inquilinos = db.session.query(
        func.count(func.distinct(DetalleAlquilerSemanal.inquilino_id))
    ).filter_by(semana_alquiler_id=semana_id).scalar() + 1
    
    semana.usuario_actualizo_id = current_user.id
    semana.fecha_hora_actualizo = datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Alquiler agregado exitosamente',
        'detalle_id': detalle.id
    })
        
    #except Exception as e:
    #    db.session.rollback()
    #    return jsonify({'success': False, 'message': str(e)}), 500


# ==========================================
# ELIMINAR DETALLE DE SEMANA
# ==========================================

@alquileres_bp.route('/alquiler/detalles/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_detalle(id):
    """Elimina un detalle de alquiler y LIBERA el vehículo"""
    
    try:
        detalle = DetalleAlquilerSemanal.query.get_or_404(id)
        semana_id = detalle.semana_alquiler_id
        vehiculo_id = detalle.vehiculo_id
        
        # Get semana
        semana = SemanaAlquiler.query.get(semana_id)
        
        if semana.estado != 'abierta':
            return jsonify({
                'success': False,
                'message': 'No se puede eliminar detalles de una semana cerrada'
            }), 400
        
        # ✅ LIBERAR VEHÍCULO (marcar como disponible)
        vehiculo = Vehiculo.query.get(vehiculo_id)
        if vehiculo:
            # Verificar que no tenga otros alquileres activos
            otros_alquileres = DetalleAlquilerSemanal.query.filter(
                and_(
                    DetalleAlquilerSemanal.vehiculo_id == vehiculo_id,
                    DetalleAlquilerSemanal.id != id,
                    DetalleAlquilerSemanal.semana_alquiler_id != semana_id
                )
            ).count()
            
            if otros_alquileres == 0:
                vehiculo.disponible = True
                vehiculo.usuario_actualizo_id = current_user.id
                vehiculo.fecha_hora_actualizo = datetime.utcnow()
        
        # Update semana totals
        semana.total_vehiculos = max(0, semana.total_vehiculos - 1)
        semana.ingreso_total = float(semana.ingreso_total or 0) - float(detalle.ingreso_calculado or 0)
        
        # Delete detalle
        db.session.delete(detalle)
        db.session.flush()
        
        # Recalculate propietarios and inquilinos
        semana.total_socios = db.session.query(
            func.count(func.distinct(DetalleAlquilerSemanal.propietario_id))
        ).filter_by(semana_alquiler_id=semana_id).scalar() or 0
        
        semana.total_inquilinos = db.session.query(
            func.count(func.distinct(DetalleAlquilerSemanal.inquilino_id))
        ).filter_by(semana_alquiler_id=semana_id).scalar() or 0
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Detalle eliminado y vehículo liberado exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ==========================================
# CERRAR SEMANA (CON VALIDACIÓN ADMIN)
# ==========================================
@alquileres_bp.route('/alquiler/semanas/<int:id>/cerrar', methods=['POST'])
@login_required
@admin_required
def cerrar_semana(id):
    """Cierra una semana de trabajo - SOLO ADMIN"""
    
    try:
        semana = SemanaAlquiler.query.get_or_404(id)
        
        if semana.estado != 'abierta':
            return jsonify({'success': False, 'message': 'La semana ya está cerrada'}), 400
        
        semana.estado = 'cerrada'
        semana.usuario_actualizo_id = current_user.id
        semana.fecha_hora_actualizo = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Semana cerrada exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ==========================================
# EXPORTAR A EXCEL
# ==========================================
@alquileres_bp.route('/alquiler/semanas/<int:id>/exportar-excel')
@login_required
def exportar_excel_semana(id):
    """Exporta los detalles de una semana a Excel"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        semana = SemanaAlquiler.query.get_or_404(id)
        detalles = DetalleAlquilerSemanal.query.filter_by(semana_alquiler_id=id).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Semana {semana.numero_semana}"
        
        # Header
        headers = [
            'Propietario', 'VehÃ­culo', 'Placa', 'Inquilino', 'Tel. Inquilino',
            'Semanal', 'DT', 'Ingreso', 'InversiÃ³n', 'Concepto Desc.',
            'NÃ³mina', '% Empresa', 'Deuda', 'NÃ³mina 2', 'Banco',
            'Conf. Pago', 'DT2'
        ]
        
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for detalle in detalles:
            vehiculo = Vehiculo.query.get(detalle.vehiculo_id)
            inquilino = Inquilino.query.get(detalle.inquilino_id)
            propietario = Propietario.query.get(detalle.propietario_id)
            banco = Banco.query.get(detalle.banco_id) if detalle.banco_id else None
            marca_modelo = vehiculo.marca_modelo_vehiculo if vehiculo else None
            
            ws.append([
                propietario.nombre_apellido if propietario else '',
                f"{marca_modelo.marca} {marca_modelo.modelo}" if marca_modelo else '',
                vehiculo.placa if vehiculo else '',
                inquilino.nombre_apellido if inquilino else '',
                inquilino.telefono if inquilino else '',
                float(detalle.precio_semanal),
                detalle.dias_trabajo,
                float(detalle.ingreso_calculado),
                float(detalle.inversion_mecanica or 0),
                detalle.concepto_descuento or '',
                float(detalle.nomina_empresa),
                float(detalle.porcentaje_empresa),
                float(detalle.monto_deuda or 0),
                float(detalle.nomina_final),
                banco.banco if banco else '',
                detalle.fecha_confirmacion_pago.strftime('%d/%m/%Y') if detalle.fecha_confirmacion_pago else '',
                detalle.dias_trabajo
            ])
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"semana_{semana.fecha_inicio.strftime('%Y%m%d')}_{semana.fecha_fin.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'error')
        return redirect(url_for('alquiler.index'))


# ==========================================
# BANCOS JSON (HELPER)
# ==========================================
@alquileres_bp.route('/alquiler/bancos/json')
@login_required
def bancos_json():
    """Retorna lista de bancos en JSON para los selects"""
    
    try:
        bancos = Banco.query.all()
        bancos_data = [
            {
                'id': banco.id,
                'banco': banco.banco,
                'cuenta': banco.cuenta
            }
            for banco in bancos
        ]
        
        return jsonify({'success': True, 'bancos': bancos_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ==========================================
# PORCENTAJES DE GANANCIA - CRUD
# ==========================================
@alquileres_bp.route('/porcentajes_ganancia')
@login_required
def porcentajes_ganancia():
    """Pantalla de gestiÃ³n de porcentajes de ganancia"""
    
    porcentajes = PorcentajeGanancia.query.order_by(
        PorcentajeGanancia.por_defecto.desc(),
        PorcentajeGanancia.activo.desc(),
        PorcentajeGanancia.porcentaje.asc()
    ).all()
    
    porcentajes_activos = sum(1 for p in porcentajes if p.activo)
    porcentaje_defecto = next((p for p in porcentajes if p.por_defecto), None)
    
    # Count semanas usando cada porcentaje
    semanas_con_porcentaje = SemanaAlquiler.query.count()
    
    return render_template(
        'modulos/porcentajes_ganancia.html',
        porcentajes=porcentajes,
        porcentajes_activos=porcentajes_activos,
        porcentaje_defecto=porcentaje_defecto,
        semanas_con_porcentaje=semanas_con_porcentaje
    )


@alquileres_bp.route('/alquiler/porcentajes_ganancia/crear', methods=['POST'])
@login_required
def crear_porcentaje_ganancia():
    """Crea un nuevo porcentaje de ganancia"""
    
    try:
        descripcion = request.form.get('descripcion')
        porcentaje = float(request.form.get('porcentaje'))
        activo = request.form.get('activo') == 'on'
        por_defecto = request.form.get('por_defecto') == 'on'
        
        # Si es por defecto, desmarcar otros
        if por_defecto:
            PorcentajeGanancia.query.update({'por_defecto': False})
        
        nuevo_porcentaje = PorcentajeGanancia(
            descripcion=descripcion,
            porcentaje=porcentaje,
            activo=activo,
            por_defecto=por_defecto,
            usuario_registro_id=current_user.id
        )
        
        db.session.add(nuevo_porcentaje)
        db.session.commit()
        
        flash('Porcentaje de ganancia creado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear porcentaje: {str(e)}', 'error')
    
    return redirect(url_for('alquiler.porcentajes_ganancia'))


@alquileres_bp.route('/alquiler/porcentajes_ganancia/<int:id>/json')
@login_required
def ver_porcentaje_json(id):
    """Retorna un porcentaje en JSON"""
    
    try:
        porcentaje = PorcentajeGanancia.query.get_or_404(id)
        
        return jsonify({
            'success': True,
            'data': {
                'id': porcentaje.id,
                'descripcion': porcentaje.descripcion,
                'porcentaje': float(porcentaje.porcentaje),
                'activo': porcentaje.activo,
                'por_defecto': porcentaje.por_defecto
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 404


@alquileres_bp.route('/alquiler/porcentajes_ganancia/<int:id>/editar', methods=['POST'])
@login_required
def editar_porcentaje_ganancia(id):
    """Edita un porcentaje de ganancia"""
    
    try:
        porcentaje = PorcentajeGanancia.query.get_or_404(id)
        
        porcentaje.descripcion = request.form.get('descripcion')
        porcentaje.porcentaje = float(request.form.get('porcentaje'))
        porcentaje.activo = request.form.get('activo') == 'on'
        por_defecto = request.form.get('por_defecto') == 'on'
        
        if por_defecto and not porcentaje.por_defecto:
            PorcentajeGanancia.query.update({'por_defecto': False})
            porcentaje.por_defecto = True
        elif not por_defecto:
            porcentaje.por_defecto = False
        
        porcentaje.usuario_actualizo_id = current_user.id
        
        db.session.commit()
        
        flash('Porcentaje actualizado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar porcentaje: {str(e)}', 'error')
    
    return redirect(url_for('alquiler.porcentajes_ganancia'))


@alquileres_bp.route('/alquiler/porcentajes_ganancia/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_porcentaje_ganancia(id):
    """Elimina un porcentaje de ganancia"""
    
    try:
        porcentaje = PorcentajeGanancia.query.get_or_404(id)
        
        # Check if in use
        if porcentaje.semanas_alquiler.count() > 0:
            flash('No se puede eliminar: el porcentaje estÃ¡ en uso', 'error')
            return redirect(url_for('alquiler.porcentajes_ganancia'))
        
        db.session.delete(porcentaje)
        flash('Porcentaje eliminado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar porcentaje: {str(e)}', 'error')
    
    return redirect(url_for('alquiler.porcentajes_ganancia'))
        
# ==========================================
# CONFIGURACIÓN MORA / INCAUTACIÓN
# ==========================================
@alquileres_bp.route('/catalogos/configuracion-alquiler', methods=['GET'])
@login_required
def configuracion_alquiler():
    """Renderiza la página de configuración de alquileres"""
    try:
        config = ConfiguracionAlquiler.query.first()
        
        # Si no existe, crear valores por defecto
        if not config:
            config = ConfiguracionAlquiler(
                dias_incautacion=3,
                porcentaje_mora=5.0
            )
            db.session.add(config)
            db.session.commit()
            
        return render_template('catalogos/configuracion_alquiler.html', config=config)
    except Exception as e:
        flash(f'Error al cargar configuración: {str(e)}', 'error')
        return redirect(url_for('alquiler.index'))

@alquileres_bp.route('/catalogos/configuracion-alquiler/guardar', methods=['POST'])
@login_required
@admin_required
def guardar_configuracion():
    """Guarda la configuración de alquileres desde el formulario"""
    try:
        dias = int(request.form.get('dias_incautacion'))
        mora = float(request.form.get('porcentaje_mora'))
        costo_lavado = float(request.form.get('costo_lavado', 350.00))
        
        config = ConfiguracionAlquiler.query.first()
        if not config:
            config = ConfiguracionAlquiler()
            db.session.add(config)
            
        config.dias_incautacion = dias
        config.porcentaje_mora = mora
        config.costo_lavado = costo_lavado
        config.usuario_actualizo_id = current_user.id
        config.fecha_hora_actualizo = datetime.utcnow()
        
        db.session.commit()
        
        flash('Configuración guardada correctamente', 'success')
        return redirect(url_for('alquiler.configuracion_alquiler'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al guardar configuración: {str(e)}', 'error')
        return redirect(url_for('alquiler.configuracion_alquiler'))


# ==========================================
# CORRECCIÓN COMPLETA: DISPONIBLES PARA ALQUILER
# Con debug y manejo correcto del campo disponible encriptado
# ==========================================

@alquileres_bp.route('/alquiler/semanas/<int:id>/disponibles')
@login_required
def disponibles_para_alquiler(id):
    """Retorna vehículos e inquilinos disponibles para agregar a la semana"""
    
    try:
        semana = SemanaAlquiler.query.get_or_404(id)
        
        print(f"🔍 DEBUG: Buscando disponibles para semana {id}")
        print(f"   Rango: {semana.fecha_inicio} a {semana.fecha_fin}")
        
        # ==========================================
        # VEHÍCULOS DISPONIBLES
        # ==========================================
        
        # 1. Vehículos ya en esta semana
        vehiculos_en_semana = db.session.query(
            DetalleAlquilerSemanal.vehiculo_id
        ).filter_by(semana_alquiler_id=id).all()
        vehiculos_ids_en_semana = [v[0] for v in vehiculos_en_semana]
        
        print(f"   Vehículos ya en semana: {vehiculos_ids_en_semana}")
        
        # 2. Vehículos con alquiler activo en el rango de fechas
        vehiculos_con_alquiler_activo = db.session.query(Alquiler.vehiculo_id).filter(
            and_(
                Alquiler.fecha_alquiler_inicio <= semana.fecha_fin,
                Alquiler.fecha_alquiler_fin >= semana.fecha_inicio
            )
        ).all()
        vehiculos_ids_con_alquiler = [v[0] for v in vehiculos_con_alquiler_activo]
        
        print(f"   Vehículos con alquiler activo: {vehiculos_ids_con_alquiler}")
        
        # 3. ✅ CORRECCIÓN: Filtrar por CONTRATO ACTIVO y DEPOSITO
        # Requerimiento: Solo incluir vehículos que tienen un contrato activo y depósito OK
        todos_vehiculos = Vehiculo.query \
            .join(Contrato, Vehiculo.contratos) \
            .join(Deposito, Contrato.depositos) \
            .filter(
                and_(
                    ~Vehiculo.id.in_(vehiculos_ids_en_semana),
                    Contrato.estado == 'activo',
                    or_(
                        Deposito.estado == 'parcial',
                        Deposito.estado == 'completado'
                    )
                )
            ).all()

        vehiculos_disponibles = todos_vehiculos
        print(f"   Vehículos disponibles con contrato activo: {len(vehiculos_disponibles)}")
        
        print(f"   Vehículos realmente disponibles: {len(vehiculos_disponibles)}")
        
        # ==========================================
        # INQUILINOS DISPONIBLES
        # ==========================================
        
        # 1. Inquilinos ya en esta semana
        inquilinos_en_semana = db.session.query(
            DetalleAlquilerSemanal.inquilino_id
        ).filter_by(semana_alquiler_id=id).all()
        inquilinos_ids_en_semana = [i[0] for i in inquilinos_en_semana]
        
        print(f"   Inquilinos ya en semana: {inquilinos_ids_en_semana}")
        
        # 2. Inquilinos con alquiler activo en el rango
        inquilinos_con_alquiler_activo = db.session.query(Alquiler.inquilino_id).filter(
            and_(
                Alquiler.fecha_alquiler_inicio <= semana.fecha_fin,
                Alquiler.fecha_alquiler_fin >= semana.fecha_inicio
            )
        ).all()
        inquilinos_ids_con_alquiler = [i[0] for i in inquilinos_con_alquiler_activo]
        
        print(f"   Inquilinos con alquiler activo: {inquilinos_ids_con_alquiler}")
        
        # 3. ✅ Inquilinos realmente disponibles (sin duplicados)
        # Requerimiento: Solo inquilinos con CONTRATO ACTIVO y DEPOSITO (Parcial o Completado)
        inquilinos_disponibles_query = Inquilino.query \
            .join(Contrato, Inquilino.contratos) \
            .join(Deposito, Contrato.depositos) \
            .filter(
                and_(
                    ~Inquilino.id.in_(inquilinos_ids_en_semana),
                    ~Inquilino.id.in_(inquilinos_ids_con_alquiler),
                    Contrato.estado == 'activo',
                    or_(
                        Deposito.estado == 'parcial',
                        Deposito.estado == 'completado'
                    )
                )
            )
        
        inquilinos_disponibles = inquilinos_disponibles_query.all()
        
        # 4. 🔴 Asegurar que los inquilinos asociados a los vehículos disponibles ESTÉN en la lista
        # (Aunque hayan sido filtrados por alguna razón, si el vehículo está libre, su dueño debería poder seleccionarse)
        inquilinos_ids_en_lista = {i.id for i in inquilinos_disponibles}
        
        for v in vehiculos_disponibles:
             # Buscar contrato activo de este vehículo
             c = Contrato.query.filter_by(vehiculo_id=v.id, estado='activo').first()
             if c and c.inquilino_id and c.inquilino_id not in inquilinos_ids_en_lista:
                 inquilino_faltante = Inquilino.query.get(c.inquilino_id)
                 if inquilino_faltante:
                     inquilinos_disponibles.append(inquilino_faltante)
                     inquilinos_ids_en_lista.add(c.inquilino_id)
                     print(f"   ⚠️ Agregado forzosamente inquilino asociado: {inquilino_faltante.nombre_apellido}")

        print(f"   Inquilinos disponibles total: {len(inquilinos_disponibles)}")
        
        # ==========================================
        # PREPARAR DATOS PARA JSON
        # ==========================================
        
        vehiculos_data = []
        for vehiculo in vehiculos_disponibles:
            try:
                # Usar relationship correctamente
                marca_modelo = vehiculo.marca_modelo if vehiculo.marca_modelo else None
                propietario = Propietario.query.get(vehiculo.propietario_id) if vehiculo.propietario_id else None
                
                # Obtener contrato activo para este vehículo para saber el inquilino
                contrato_activo = Contrato.query.filter_by(vehiculo_id=vehiculo.id, estado='activo').first()
                if contrato_activo:
                    inquilino_asociado_id = contrato_activo.inquilino_id
                    inquilino_asoc = Inquilino.query.get(inquilino_asociado_id)
                    inquilino_nombre = inquilino_asoc.nombre_apellido if inquilino_asoc else "Desconocido"
                else:
                    inquilino_asociado_id = None
                    inquilino_nombre = None
                
                vehiculo_dict = {
                    'id': vehiculo.id,
                    'placa': vehiculo.placa if vehiculo.placa else 'N/A',
                    'marca': marca_modelo.marca if marca_modelo else 'N/A',
                    'modelo': marca_modelo.modelo if marca_modelo else 'N/A',
                    'ano': vehiculo.ano if vehiculo.ano else '',
                    'color': vehiculo.color if vehiculo.color else '',
                    'precio_semanal': float(vehiculo.precio_semanal) if vehiculo.precio_semanal else 0,
                    'propietario_id': vehiculo.propietario_id,
                    'propietario_nombre': propietario.nombre_apellido if propietario else '',
                    'inquilino_asociado_id': inquilino_asociado_id, # ID para auto-selección
                    'inquilino_nombre': inquilino_nombre, # Nombre para display
                    'disponible': True  # Ya filtrados
                }
                
                vehiculos_data.append(vehiculo_dict)
                
            except Exception as e:
                print(f"❌ Error procesando vehículo {vehiculo.id}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        # ✅ Ordenar por placa
        vehiculos_data.sort(key=lambda x: x['placa'])
        
        inquilinos_data = []
        for inquilino in inquilinos_disponibles:
            try:
                inquilino_dict = {
                    'id': inquilino.id,
                    'nombre_apellido': inquilino.nombre_apellido if inquilino.nombre_apellido else 'Sin nombre',
                    'telefono': inquilino.telefono if inquilino.telefono else '',
                    'cedula': inquilino.cedula if inquilino.cedula else ''
                }
                
                inquilinos_data.append(inquilino_dict)
                
            except Exception as e:
                print(f"❌ Error procesando inquilino {inquilino.id}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        # ✅ Ordenar por nombre y ELIMINAR DUPLICADOS
        inquilinos_data.sort(key=lambda x: x['nombre_apellido'])
        
        # ✅ ELIMINAR DUPLICADOS por ID
        inquilinos_unicos = []
        ids_vistos = set()
        for inq in inquilinos_data:
            if inq['id'] not in ids_vistos:
                inquilinos_unicos.append(inq)
                ids_vistos.add(inq['id'])
        
        print(f"✅ Respuesta: {len(vehiculos_data)} vehículos, {len(inquilinos_unicos)} inquilinos")
        
        return jsonify({
            'success': True,
            'vehiculos': vehiculos_data,
            'inquilinos': inquilinos_unicos,
            'debug': {
                'semana_id': id,
                'total_vehiculos': len(vehiculos_data),
                'total_inquilinos': len(inquilinos_unicos),
                'vehiculos_en_semana': len(vehiculos_ids_en_semana),
                'inquilinos_en_semana': len(inquilinos_ids_en_semana)
            }
        })
        
    except Exception as e:
        print(f"❌ ERROR CRÍTICO en disponibles_para_alquiler: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==========================================
# CORRECCIÓN 5: CREAR INVERSIÓN MECÁNICA (CON TODOS LOS CAMPOS REQUERIDOS)
# ==========================================
@alquileres_bp.route('/alquiler/inversiones/crear', methods=['POST'])
@login_required
def crear_inversion_mecanica():
    """Crea una inversión mecánica asociada a un alquiler"""
    
    try:
        data = request.get_json()
        
        detalle_id = int(data.get('detalle_id'))
        mecanico_id = int(data.get('mecanico_id'))
        tipo_trabajo_id = int(data.get('tipo_trabajo_id'))
        tipo_inversion = data.get('tipo_inversion')
        descripcion = data.get('descripcion')
        costo = float(data.get('costo'))
        
        # Validate detalle exists
        detalle = DetalleAlquilerSemanal.query.get_or_404(detalle_id)
        semana = SemanaAlquiler.query.get(detalle.semana_alquiler_id)
        
        # ✅ CORRECCIÓN: Usar fecha de la SEMANA, no date.today()
        # Esto asegura que la inversión esté dentro del rango correcto
        fecha_trabajo = semana.fecha_inicio
        
        print(f"📅 Creando inversión con fecha de semana: {fecha_trabajo}")
        print(f"   Rango semana: {semana.fecha_inicio} - {semana.fecha_fin}")
        
        trabajo = TrabajoVehiculo(
            vehiculo_id=detalle.vehiculo_id,
            mecanico_id=mecanico_id,
            tipo_trabajo_id=tipo_trabajo_id,
            fecha_inicio=fecha_trabajo,  # Fecha de inicio de semana
            fecha_fin=fecha_trabajo,     # Fecha de inicio de semana
            descripcion=descripcion,
            costo=costo,
            estado='completado',
            notas=f'Tipo: {tipo_inversion}',
            usuario_registro_id=current_user.id,
            usuario_actualizo_id=current_user.id
        )
        
        db.session.add(trabajo)
        db.session.flush()
        
        # ✅ Calcular suma TOTAL de inversiones del vehículo en esta semana
        total_inversiones = db.session.query(
            func.sum(TrabajoVehiculo.costo)
        ).filter(
            and_(
                TrabajoVehiculo.vehiculo_id == detalle.vehiculo_id,
                TrabajoVehiculo.fecha_inicio >= semana.fecha_inicio,
                TrabajoVehiculo.fecha_inicio <= semana.fecha_fin
            )
        ).scalar() or 0
        
        print(f"✅ Inversión creada ID {trabajo.id}: ${costo}")
        print(f"   Total acumulado del vehículo: ${total_inversiones}")
        
        # Actualizar campo de inversión mecánica
        detalle.inversion_mecanica = total_inversiones
        
        # Actualizar concepto
        if not detalle.concepto_inversion:
            detalle.concepto_inversion = descripcion
        else:
            if descripcion not in detalle.concepto_inversion:
                detalle.concepto_inversion += f"; {descripcion}"
        
        detalle.usuario_actualizo_id = current_user.id
        detalle.fecha_hora_actualizo = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Inversión registrada exitosamente',
            'trabajo_id': trabajo.id,
            'total_inversion': float(total_inversiones)
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al crear inversión: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==========================================
# ELIMINAR SEMANA (CON VALIDACIÓN ADMIN)
# ==========================================
@alquileres_bp.route('/alquiler/semanas/<int:id>/eliminar', methods=['POST'])
@login_required
@admin_required
def eliminar_semana(id):
    """Elimina una semana - SOLO ADMIN"""
    
    try:
        semana = SemanaAlquiler.query.get_or_404(id)
        
        # Count detalles
        detalles_count = DetalleAlquilerSemanal.query.filter_by(
            semana_alquiler_id=id
        ).count()
        
        # Delete detalles first (si hay)
        if detalles_count > 0:
            DetalleAlquilerSemanal.query.filter_by(
                semana_alquiler_id=id
            ).delete()
        
        # Delete semana
        db.session.delete(semana)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Semana eliminada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ==========================================
# OBTENER INVERSIONES DE DETALLE (NUEVO)
# ==========================================
@alquileres_bp.route('/alquiler/detalles/<int:id>/inversiones')
@login_required
def obtener_inversiones_detalle(id):
    """Obtiene las inversiones de un detalle de alquiler"""
    
    try:
        detalle = DetalleAlquilerSemanal.query.get_or_404(id)
        semana = SemanaAlquiler.query.get(detalle.semana_alquiler_id)
        
        print(f"🔍 Buscando inversiones:")
        print(f"   Vehículo ID: {detalle.vehiculo_id}")
        print(f"   Rango: {semana.fecha_inicio} - {semana.fecha_fin}")
        
        # Get trabajos del vehículo en el rango de la semana
        trabajos = TrabajoVehiculo.query.filter(
            and_(
                TrabajoVehiculo.vehiculo_id == detalle.vehiculo_id,
                TrabajoVehiculo.fecha_inicio >= semana.fecha_inicio,
                TrabajoVehiculo.fecha_inicio <= semana.fecha_fin
            )
        ).order_by(TrabajoVehiculo.fecha_inicio.desc(), TrabajoVehiculo.id.desc()).all()
        
        print(f"   Encontrados: {len(trabajos)} trabajos")
        
        inversiones_data = []
        for trabajo in trabajos:
            mecanico = Mecanico.query.get(trabajo.mecanico_id)
            tipo_trabajo = TipoTrabajo.query.get(trabajo.tipo_trabajo_id)
            
            print(f"      - ID {trabajo.id}: ${trabajo.costo} ({trabajo.fecha_inicio})")
            
            nombre_tipo = 'N/A'
            if tipo_trabajo:
                nombre_tipo = tipo_trabajo.nombre
            elif trabajo.descripcion == 'Lavado Automático al Ingreso':
                nombre_tipo = 'Lavado'
            
            tipo_inversion = 'N/A'
            if trabajo.notas and trabajo.notas.startswith('Tipo: '):
                tipo_inversion = trabajo.notas.replace('Tipo: ', '').split('|')[0].strip()
            elif trabajo.descripcion == 'Lavado Automático al Ingreso':
                tipo_inversion = 'Mantenimiento'
            
            inversiones_data.append({
                'id': trabajo.id,
                'fecha': trabajo.fecha_inicio.strftime('%d/%m/%Y'),
                'mecanico': mecanico.nombre if mecanico else 'N/A',
                'tipo_trabajo': nombre_tipo,
                'descripcion': trabajo.descripcion,
                'costo': float(trabajo.costo),
                'tipo_inversion': tipo_inversion
            })
        
        total_inversiones = sum(float(t.costo) for t in trabajos)
        
        print(f"✅ Total: ${total_inversiones}")
        
        return jsonify({
            'success': True,
            'inversiones': inversiones_data,
            'total': total_inversiones
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo inversiones: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==========================================
# EDITAR DETALLE COMPLETO (CON VEHÃCULO E INQUILINO)
# ==========================================
@alquileres_bp.route('/alquiler/detalles/<int:id>/editar_completo', methods=['POST'])
@login_required
def editar_detalle_completo(id):
    """Edita un detalle completo incluyendo vehÃ­culo e inquilino"""
    
    try:
        detalle = DetalleAlquilerSemanal.query.get_or_404(id)
        semana = SemanaAlquiler.query.get(detalle.semana_alquiler_id)
        
        if semana.estado != 'abierta':
            return jsonify({
                'success': False,
                'message': 'No se puede editar una semana cerrada'
            }), 400
        
        data = request.get_json()
        
        # Get new values
        nuevo_vehiculo_id = int(data.get('vehiculo_id'))
        nuevo_inquilino_id = int(data.get('inquilino_id'))
        
        # Validate vehÃ­culo change
        if nuevo_vehiculo_id != detalle.vehiculo_id:
            # Check if new vehiculo is already in this semana
            existing = DetalleAlquilerSemanal.query.filter(
                and_(
                    DetalleAlquilerSemanal.semana_alquiler_id == detalle.semana_alquiler_id,
                    DetalleAlquilerSemanal.vehiculo_id == nuevo_vehiculo_id,
                    DetalleAlquilerSemanal.id != id
                )
            ).first()
            
            if existing:
                return jsonify({
                    'success': False,
                    'message': 'Este vehÃ­culo ya estÃ¡ asignado en esta semana'
                }), 400
        
        # Validate inquilino change
        if nuevo_inquilino_id != detalle.inquilino_id:
            # Check if new inquilino is already in this semana
            existing_inq = DetalleAlquilerSemanal.query.filter(
                and_(
                    DetalleAlquilerSemanal.semana_alquiler_id == detalle.semana_alquiler_id,
                    DetalleAlquilerSemanal.inquilino_id == nuevo_inquilino_id,
                    DetalleAlquilerSemanal.id != id
                )
            ).first()
            
            if existing_inq:
                return jsonify({
                    'success': False,
                    'message': 'Este inquilino ya tiene un vehÃ­culo asignado en esta semana'
                }), 400
        
        # Get vehÃ­culo data
        vehiculo = Vehiculo.query.get_or_404(nuevo_vehiculo_id)
        
        # Update detalle
        detalle.vehiculo_id = nuevo_vehiculo_id
        detalle.inquilino_id = nuevo_inquilino_id
        detalle.propietario_id = vehiculo.propietario_id
        detalle.precio_semanal = data.get('precio_semanal', vehiculo.precio_semanal)
        detalle.dias_trabajo = int(data.get('dias_trabajo') or 7)
        detalle.inversion_mecanica = float(data.get('inversion_mecanica') or 0)
        detalle.concepto_inversion = data.get('concepto_inversion', '')
        detalle.monto_descuento = float(data.get('monto_descuento') or 0)
        detalle.concepto_descuento = data.get('concepto_descuento', '')
        detalle.monto_deuda = float(data.get('monto_deuda') or 0)
        detalle.banco_id = data.get('banco_id') if data.get('banco_id') else None
        
        if data.get('fecha_confirmacion_pago'):
            detalle.fecha_confirmacion_pago = datetime.strptime(
                data.get('fecha_confirmacion_pago'), '%Y-%m-%d'
            ).date()
        
        detalle.pago_confirmado = data.get('pago_confirmado', False)
        detalle.notas = data.get('notas', '')
        
        # Recalculate
        precio_diario = float(detalle.precio_semanal) / 7.0
        detalle.ingreso_calculado = int(round(precio_diario * float(detalle.dias_trabajo)))
        
        # Nomina Calculada: (Ingreso * 0.9) - Inversion
        detalle.nomina_empresa = (detalle.ingreso_calculado * 0.9) - float(detalle.inversion_mecanica or 0)
        
        detalle.nomina_final = detalle.ingreso_calculado + float(detalle.monto_deuda or 0)
        
        detalle.usuario_actualizo_id = current_user.id
        detalle.fecha_hora_actualizo = datetime.utcnow()
        
        # Update semana totals
        semana.usuario_actualizo_id = current_user.id
        semana.fecha_hora_actualizo = datetime.utcnow()
        
        # Recalculate totals
        detalles = DetalleAlquilerSemanal.query.filter_by(
            semana_alquiler_id=detalle.semana_alquiler_id
        ).all()
        
        semana.ingreso_total = sum(float(d.ingreso_calculado) for d in detalles)
        semana.total_vehiculos = len(detalles)
        
        # Recalculate unique counts
        semana.total_socios = db.session.query(
            func.count(func.distinct(DetalleAlquilerSemanal.propietario_id))
        ).filter_by(semana_alquiler_id=detalle.semana_alquiler_id).scalar()
        
        semana.total_inquilinos = db.session.query(
            func.count(func.distinct(DetalleAlquilerSemanal.inquilino_id))
        ).filter_by(semana_alquiler_id=detalle.semana_alquiler_id).scalar()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Detalle actualizado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==========================================
# CONFIRMAR PAGO (MODAL DEDICADO)
# ==========================================
@alquileres_bp.route('/alquiler/detalles/<int:id>/confirmar_pago', methods=['POST'])
@login_required
def confirmar_pago_detalle(id):
    """Confirmar pago de un detalle con banco, fecha y comprobante"""
    
    try:
        detalle = DetalleAlquilerSemanal.query.get_or_404(id)
        semana = SemanaAlquiler.query.get(detalle.semana_alquiler_id)
        
        if semana.estado != 'abierta' and current_user.rol != 'admin':
            return jsonify({
                'success': False,
                'message': 'No se puede confirmar pago en una semana cerrada (requiere admin)'
            }), 400
            
        banco_id = request.form.get('banco_id')
        fecha_pago = request.form.get('fecha_confirmacion_pago')
        pago_confirmado = request.form.get('pago_confirmado') == 'true'
        
        if banco_id == '0':
            detalle.banco_id = None
        elif banco_id:
            detalle.banco_id = int(banco_id)
        if fecha_pago:
            try:
                detalle.fecha_confirmacion_pago = datetime.strptime(fecha_pago, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                detalle.fecha_confirmacion_pago = date.today()
        else:
            detalle.fecha_confirmacion_pago = date.today()
            
        detalle.pago_confirmado = pago_confirmado
        
        # Manejo de archivo (comprobante)
        eliminar_comprobante = request.form.get('eliminar_comprobante') == 'true'
        
        if eliminar_comprobante:
            detalle.comprobante_pago_path = None

        if 'comprobante' in request.files:
            file = request.files['comprobante']
            if file and file.filename != '':
                # Crear directorio si no existe
                upload_folder = os.path.join(current_app.static_folder, 'uploads', 'pagos_alquiler')
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)
                
                filename = secure_filename(f"pago_{id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                file_path = os.path.join(upload_folder, filename)
                file.save(file_path)
                
                # Guardar ruta relativa en el modelo (sobrescribe cualquier flag de eliminación)
                detalle.comprobante_pago_path = os.path.join('uploads', 'pagos_alquiler', filename)
        
        detalle.usuario_actualizo_id = current_user.id
        detalle.fecha_hora_actualizo = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': 'Pago confirmado correctamente',
            'pago_confirmado': detalle.pago_confirmado,
            'banco': detalle.banco.banco if detalle.banco else 'N/A'
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error al confirmar pago:\n{error_details}")
        return jsonify({
            'success': False, 
            'message': str(e),
            'debug': error_details if current_app.debug else None
        }), 500
"""
En tu app/__init__.py, agregar:

from app.routes.alquileres_routes import alquileres_bp
app.register_blueprint(alquileres_bp)
"""